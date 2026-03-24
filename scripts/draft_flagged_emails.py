"""
draft_flagged_emails.py
-----------------------
Reads daily report Excel files and drafts Outlook emails to pods.

Port Review  → Port_Review_*.xlsx                      → "Port Review Master Table" sheet
               Col T (NET_CASH): flagged when cell background is red

TD Report    → Daily_Tracking_Difference_*.xlsx        → "Daily Tracking Difference" sheet
               Col O (Flag) contains "Flag"

Corp Actions → New Corporate Actions Tracker*.xlsx     → "Summary" sheet
               Opening Items (col E) is a positive number
"""

import glob
import os
import re
import win32com.client
import xlwings as xw
import openpyxl
from datetime import datetime

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
CORP_ACTIONS_FOLDER  = r"U:\new automated report emails"
CORP_ACTIONS_PATTERN = "New Corporate Actions Tracker*.xlsx"

_today_str = datetime.today().strftime("%Y%m%d")

PORT_REVIEW_PATH = rf"X:\PM & Operations\Portfolio Management\Portfolio Review\Port_Review_{_today_str}.xlsx"
TD_PATH          = rf"X:\PM & Operations\Portfolio Management\Daily Tracking Difference Report\Daily_Tracking_Difference_{_today_str}.xlsx"

PORT_SHEET      = "Port Review Master Table"
TD_SHEET        = "Daily Tracking Difference"

PORT_TICKER_COL = 1   # Column A = FUND_TICKER
PORT_FLAG_COL   = 20  # Column T = NET_CASH (red fill = flagged)

TD_TICKER_COL   = 1   # Column A = FUND
TD_FLAG_COL     = 15  # Column O = Flag

CA_SHEET        = "Blotter"
CA_HEADER_ROW   = 2   # headers on row 2
CA_DATA_START   = 3   # data from row 3
CA_KEEP_COLS    = {   # col index (1-based) → display name
    1: "Action Date",
    3: "Index Effective",
    5: "Fund",
    6: "Parent Security Name",
    7: "Ticker",
    8: "SEI ID",
}

# ──────────────────────────────────────────────────────────────────────────────

# ── POD DEFINITIONS ───────────────────────────────────────────────────────────
PODS = {
    "vanessa": {
        "display": "Vanessa",
        "to": ["vyang@globalxetfs.com"],
        "cc": ["btran@globalxetfs.com", "jqi@globalxetfs.com"],
        "tickers": {
            # Income
            "SDIV","DIV","EFAS","SRET","ALTY","QDIV","PFFD","PFFV","SPFF",
            "SDEM","FLOW","MLPA","MLPX","AUSF",
            # Covered Call
            "QYLD","XYLD","RYLD","QYLG","XYLG","DJIA","RYLG","DYLG","TYLG",
            "MLPD","EDGQ","EDGX",
            # Options
            "XRMI","XTR","XCLR","QRMI","QTR","QCLR",
        },
    },
    "wayne_syon_nam": {
        "display": "Wayne / Syon / Nam",
        "to": [
            "wxie@globalxetfs.com",
            "nto@globalxetfs.com",
            "sverma@globalxetfs.com",
            "jwallick@globalxetfs.com",
            "mkravchenko@globalxetfs.com",
        ],
        "cc": [],
        "tickers": {
            # Int'l Access (Wayne/Syon)
            "GXG","ARGT",
            # Thematic Disruptive
            "AIQ","BKCH","BOTZ","BUG","CLOU","DRIV","FINX","HERO","SNSR",
            "SOCL","VPN","ZAP","GXDW","CHPX",
            # Thematic Physical Environment
            "AQWA","CTEC","HYDR","KROP","PAVE","RNRG","SHLD","CEFA","IPAV",
            # Broadmarket
            "CATH","KRMA","RSSL","FLAG","EGLE","GURU","CHRI","GXLC",
            # Thematic People & Demo
            "AGNG","EBIZ","EDOC","GNOM","MILN",
            # Int'l Access (Nam)
            "CHIQ","GREK","NORW","DAX","ASEA","VNAM","AUAU",
            # Commodities
            "LIT","URA","COPX","GOEX","DMAT","SIL","LNGX",
        },
    },
    "sandy": {
        "display": "Sandy",
        "to": ["slu@globalxetfs.com"],
        "cc": ["czhao@globalxetfs.com", "kszeto@globalxetfs.com"],
        "tickers": {
            # Fixed Income
            "ONOF","IRVH","CLIP","SLDR","MLDR","LLDR",
            "ZCBA","ZCBB","ZCBC","ZCBE","ZCBF","ZCBG",
            # Derivatives
            "BITS","BTRN","BCCC","TLTX","COMD",
            # Quant Solutions
            "GXPT","GXPC","GXPD","GXPE","GXPS",
            # Digital Assets
            "BT0X GR","ET0X GR","LI0X GR","UNIX GR","AVMX GR",
        },
    },
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def _latest_file(folder: str, pattern: str) -> str:
    """Find the most recently modified file matching pattern in folder."""
    matches = glob.glob(os.path.join(folder, pattern))
    if not matches:
        raise FileNotFoundError(f"No file found matching: {pattern}")
    return max(matches, key=os.path.getmtime)


_VANESSA_ALIASES = {"X FUNDS", "Q FUNDS"}


def _get_pod(ticker: str):
    t = str(ticker).strip().upper()
    if t in _VANESSA_ALIASES:
        return "vanessa"
    for pod_name, pod in PODS.items():
        if t in pod["tickers"]:
            return pod_name
    return None


def _excel_color_is_red(raw_color: int) -> bool:
    """
    Excel COM Interior.Color is stored as BGR integer.
    Any non-white background in column T is treated as flagged because the only
    CF rule on that column applies a red-derived fill.
    """
    white = 16777215   # RGB(255,255,255)
    no_fill = -4142    # xlNone
    return raw_color not in (white, no_fill)


def _parse_bps(val):
    """Parse a value like '3 bps', '-5 bps', or a plain number. Returns float or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        pass
    m = re.search(r"-?\d+(\.\d+)?", str(val))
    return float(m.group()) if m else None


_PCT_COLS = {
    "ERROR_CHECK",
    "CUSTODY_CASH_USD_ADJ",
    "ACTUAL_CASH",
    "ACCRUED_CASH",
    "NET_CASH",
    "NET_CASH_FUTURES_ADJ",
}

_DOLLAR_COLS = {"ASSETS", "AUM"}

_CASH_COLS = {
    "CUSTODY_CASH_USD_ADJ",
    "ACTUAL_CASH",
    "ACCRUED_CASH",
    "NET_CASH",
    "NET_CASH_FUTURES_ADJ",
}

_CASH_WIDTH = "min-width:90px;width:90px;"

_ATTR_BPS_COLS = {
    "CASH DRAG",
    "SEC LENDING",
    "RECLAIMS",
    "DIV ADJ",
    "REALIZED G/L",
    "TRANS COSTS",
    "FAIR VALUE",
    "ACTIVE WT",
    "WH DISC",
    "FEES",
    "FUTURES",
    "INDIA TAX",
    "EXPLAINED TD",
    "REALIZED TD",
    "UNEXPLAINED TD",
}

_ATTR_SKIP_COLS = {"DATE", "TICKER"}

_ATTR_COL_NAMES = {
    "CASH_DRAG":           "Cash Drag",
    "SECURITIES_LENDING":  "Sec Lending",
    "RECLAIMS":            "Reclaims",
    "DIV_ADJ":             "Div Adj",
    "REALIZED_GAIN_LOSS":  "Realized G/L",
    "TRANSACTION_COSTS":   "Trans Costs",
    "PX_FV":               "Fair Value",
    "ACTIVE_WEIGHT":       "Active Wt",
    "WH_DISCREPANCY":      "WH Disc",
    "FEES":                "Fees",
    "FUTURES":             "Futures",
    "INDIA_TAX":           "India Tax",
    "EXPLAINED_TD":        "Explained TD",
    "REALIZED_TD":         "Realized TD",
    "UNEXPLAINED_TD":      "Unexplained TD",
    "Explanation":         "Explanation",
    "ASSETS":              "AUM",
    "NAV":                 "NAV",
}


def _fmt(v, header=None):
    """Format a cell value based on its column header."""
    h = str(header).strip().upper() if header else ""

    if h in _DOLLAR_COLS:
        if v is None or v == "":
            return ""
        return f"${float(v):,.0f}"

    if h in _PCT_COLS:
        if v is None or v == "":
            return ""
        return f"{float(v) * 100:.2f}%"

    if isinstance(header, datetime) or h in _ATTR_BPS_COLS:
        if v is None or v == "":
            return "0.0bp"
        return f"{float(v):.1f}bp"

    if isinstance(v, (int, float)):
        return f"{float(v):.2f}"

    if isinstance(v, datetime):
        return v.strftime("%m/%d/%Y")

    return v if v is not None else ""


def _is_numeric(v) -> bool:
    """True if value should be right-aligned."""
    return isinstance(v, (int, float))


def _html_table(headers: list, rows: list, keep_cols: list = None, uniform_width: str = None) -> str:
    """
    Render an HTML table.
    If there is an Explanation column, show it as a full-width row
    underneath each data row instead of as the last column.
    Date headers stay blue; date data cells are light beige.
    """
    def _is_date_col(h):
        return isinstance(h, datetime)

    if keep_cols is not None:
        headers = [headers[i] for i in keep_cols]
        rows = [[row[i] for i in keep_cols] for row in rows]

    def _is_explanation_col(h):
        return str(h).strip().upper() == "EXPLANATION"

    def _px_to_int(px_val, default=70):
        if not px_val:
            return default
        try:
            return int(str(px_val).replace("px", "").strip())
        except Exception:
            return default

    normal_width = _px_to_int(uniform_width, 70)

    expl_idx = None
    for i, h in enumerate(headers):
        if _is_explanation_col(h):
            expl_idx = i
            break

    if expl_idx is not None:
        table_headers = headers[:expl_idx] + headers[expl_idx + 1:]
        table_rows = []
        explanations = []
        for row in rows:
            explanations.append(row[expl_idx])
            table_rows.append(row[:expl_idx] + row[expl_idx + 1:])
    else:
        table_headers = headers
        table_rows = rows
        explanations = [None] * len(rows)

    def _th_style(h):
        h_up = str(h).strip().upper() if not isinstance(h, datetime) else ""
        style = (
            "border:1px solid #bbb;"
            "padding:5px 6px;"
            "background:#dce6f1;"
            "font-family:Calibri,sans-serif;"
            "font-size:10pt;"
            "font-weight:bold;"
            "vertical-align:top;"
            "line-height:14px;"
            "mso-line-height-rule:exactly;"
            "white-space:normal;"
        )
        if h_up in _PCT_COLS or h_up in _DOLLAR_COLS or h_up in _ATTR_BPS_COLS or isinstance(h, datetime):
            style += "text-align:right;"
        else:
            style += "text-align:center;"
        return style

    def _td_style(h, v, bg):
        cell_bg = "#f8f1e4" if _is_date_col(h) else bg
        style = (
            f"border:1px solid #bbb;"
            f"padding:4px 6px;"
            f"background:{cell_bg};"
            "font-family:Calibri,sans-serif;"
            "font-size:10pt;"
            "vertical-align:top;"
            "line-height:14px;"
            "mso-line-height-rule:exactly;"
            "white-space:nowrap;"
        )
        if _is_numeric(v) or isinstance(h, datetime):
            style += "text-align:right;"
        else:
            style += "text-align:left;"
        return style

    def _expl_style(bg):
        return (
            f"border:1px solid #bbb;"
            f"border-top:none;"
            f"padding:6px 8px;"
            f"background:{bg};"
            "font-family:Calibri,sans-serif;"
            "font-size:10pt;"
            "line-height:14px;"
            "mso-line-height-rule:exactly;"
            "white-space:normal;"
            "text-align:left;"
        )

    total_width = len(table_headers) * normal_width

    th = "".join(
        f"<th width='{normal_width}' style='{_th_style(h)}'>"
        f"{h.strftime('%m/%d/%Y') if isinstance(h, datetime) else (h or '')}"
        f"</th>"
        for h in table_headers
    )

    body = ""
    for i, row in enumerate(table_rows):
        bg = "#ffffff" if i % 2 == 0 else "#f5f5f5"

        tds = "".join(
            f"<td width='{normal_width}' style='{_td_style(table_headers[j], v, bg)}'>{_fmt(v, table_headers[j])}</td>"
            for j, v in enumerate(row)
        )
        body += f"<tr>{tds}</tr>"

        expl = explanations[i]
        if expl_idx is not None and expl not in (None, "", 0):
            body += (
                f"<tr>"
                f"<td colspan='{len(table_headers)}' style='{_expl_style(bg)}'>"
                f"<b>Explanation:</b> {expl}"
                f"</td>"
                f"</tr>"
            )
            body += (
                f"<tr>"
                f"<td colspan='{len(table_headers)}' style='height:10px;border:none;background:#ffffff;'></td>"
                f"</tr>"
            )

    return (
        f"<table width='{total_width}' style='border-collapse:collapse;font-family:Calibri,sans-serif;'>"
        f"<thead><tr>{th}</tr></thead>"
        f"<tbody>{body}</tbody>"
        f"</table>"
    )


def _section_header(title: str) -> str:
    return (
        f"<p style='font-family:Calibri,sans-serif;font-size:11pt;margin-bottom:4px;'>"
        f"<b>{title}</b></p>"
    )


def _draft_email(to: list, cc: list, subject: str, html_body: str):
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = "; ".join(to + cc)
    mail.Subject = subject
    mail.Display()
    signature = mail.HTMLBody
    mail.HTMLBody = html_body + signature


# ── FILE READERS ──────────────────────────────────────────────────────────────

def read_port_review():
    """
    Uses xlwings to read the actual rendered cell color in column T.
    Any non-white background = flagged.
    Returns (headers, buckets) where buckets = {pod: [row_value_lists]}.
    """
    path = PORT_REVIEW_PATH
    print(f"Port Review file : {os.path.basename(path)}")

    app = xw.App(visible=False)
    try:
        wb_xw = app.books.open(path)
        ws_xw = wb_xw.sheets[PORT_SHEET]
        wb_xw.api.Application.CalculateFull()

        last_row = ws_xw.range("A1").end("down").row
        last_col = ws_xw.used_range.last_cell.column

        data = ws_xw.range(
            ws_xw.cells(1, 1),
            ws_xw.cells(last_row, last_col)
        ).value

        flag_colors = [
            int(ws_xw.cells(r, PORT_FLAG_COL).api.DisplayFormat.Interior.Color)
            for r in range(2, last_row + 1)
        ]
    finally:
        wb_xw.close()
        app.quit()

    headers = data[0]
    buckets = {pod: [] for pod in PODS}

    for i, row_vals in enumerate(data[1:]):
        ticker = row_vals[PORT_TICKER_COL - 1]
        if not ticker:
            continue
        if _excel_color_is_red(flag_colors[i]):
            pod = _get_pod(str(ticker))
            if pod:
                buckets[pod].append(row_vals)

    return headers, buckets


def read_td_report():
    """
    Reads the Daily Tracking Difference sheet and keeps rows flagged in column O.
    Returns (headers, buckets).
    """
    path = TD_PATH
    print(f"TD Report file   : {os.path.basename(path)}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[TD_SHEET]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    buckets = {pod: [] for pod in PODS}

    for row in ws.iter_rows(min_row=2):
        ticker = row[TD_TICKER_COL - 1].value
        if not ticker:
            continue

        if str(row[TD_FLAG_COL - 1].value).strip().lower() == "flag":
            pod = _get_pod(str(ticker))
            if pod:
                buckets[pod].append([cell.value for cell in row])

    return headers, buckets


r'''
def read_corp_actions():
    """
    Reads the Blotter sheet of the Corporate Actions Tracker.
    Filters for today's Action Date and matches Fund (col 5) to pod tickers.
    Fund cells may contain multiple tickers separated by newlines — each is checked.
    Returns (headers, buckets).
    """
    path = _latest_file(CORP_ACTIONS_FOLDER, CORP_ACTIONS_PATTERN)
    print(f"Corp Actions file: {os.path.basename(path)}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[CA_SHEET]

    col_indices = sorted(CA_KEEP_COLS.keys())
    headers     = [CA_KEEP_COLS[i] for i in col_indices]
    buckets     = {pod: [] for pod in PODS}
    today_date  = datetime.today().date()

    for row in ws.iter_rows(min_row=CA_DATA_START):
        index_effective = row[2].value

        if not isinstance(index_effective, datetime) or index_effective.date() < today_date:
            continue

        fund_val = row[4].value
        if not fund_val:
            continue

        tickers = [t.strip() for t in str(fund_val).split("\n") if t.strip()]
        row_vals = [row[i - 1].value for i in col_indices]

        matched_pods = set()
        for ticker in tickers:
            pod = _get_pod(ticker)
            if pod and pod not in matched_pods:
                buckets[pod].append(row_vals)
                matched_pods.add(pod)

    return headers, buckets
'''


def read_attribution():
    """
    Reads the Attribution tab of the TD file.
    Returns (attr_headers, attr_lookup) where:
      attr_headers = list of column names (DATE and TICKER excluded)
      attr_lookup  = {ticker: [row values]} for the most recent date per ticker
    """
    wb = openpyxl.load_workbook(TD_PATH, data_only=True)
    ws = wb["Attribution"]

    all_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    keep_idx = [
        i for i, h in enumerate(all_headers)
        if str(h).strip().upper() not in _ATTR_SKIP_COLS
    ]
    attr_headers = [_ATTR_COL_NAMES.get(str(all_headers[i]), str(all_headers[i])) for i in keep_idx]

    latest = {}
    for row in ws.iter_rows(min_row=2):
        ticker = row[1].value
        date = row[0].value
        if not ticker or not isinstance(date, datetime):
            continue
        if ticker not in latest or date > latest[ticker][0]:
            latest[ticker] = (date, [row[i].value for i in keep_idx])

    attr_lookup = {t: vals for t, (_, vals) in latest.items()}
    return attr_headers, attr_lookup


# ── MAIN ──────────────────────────────────────────────────────────────────────

def _port_review_cols(headers: list) -> list:
    """
    Return indices for the selected Port Review columns, including NET_CASH_FUTURES_ADJ.
    """
    want = {
        "FUND_TICKER",
        "ATTRIBUTE",
        "ASSETS",
        "DATE",
        "ERROR_CHECK",
        "CUSTODY_CASH_USD_ADJ",
        "ACTUAL_CASH",
        "ACCRUED_CASH",
        "NET_CASH",
        "NET_CASH_FUTURES_ADJ",
    }
    return [i for i, h in enumerate(headers) if str(h).strip().upper() in want]


def _td_cols(headers: list) -> list:
    """
    Return indices for FUND (ticker) plus any column whose header is a date.
    """
    cols = []
    seen_dates = set()
    for i, h in enumerate(headers):
        if str(headers[i]).strip().upper() == "FUND":
            cols.append(i)
        elif isinstance(h, datetime) and h not in seen_dates:
            cols.append(i)
            seen_dates.add(h)
    return cols


def main():
    port_headers, port_buckets = read_port_review()
    td_headers, td_buckets = read_td_report()
    attr_headers, attr_lookup = read_attribution()
    # ca_headers, ca_buckets = read_corp_actions()

    port_cols = _port_review_cols(port_headers)
    td_cols = _td_cols(td_headers)

    td_date_headers = [td_headers[i] for i in td_cols]
    combined_td_hdrs = [
        h for h in (td_date_headers + attr_headers)
        if str(h).strip().upper() != "NAV"
    ]

    for pod in td_buckets:
        joined = []
        for row in td_buckets[pod]:
            ticker = str(row[0]).strip().upper()
            td_part = [row[i] for i in td_cols]
            raw_attr = attr_lookup.get(ticker, [None] * len(attr_headers))

            attr_vals = [
                (0.0 if v is None and i < len(attr_headers) - 1 else v)
                for i, v in enumerate(raw_attr)
            ]

            combined_row = td_part + attr_vals

            filtered_row = [
                v for h, v in zip(td_date_headers + attr_headers, combined_row)
                if str(h).strip().upper() != "NAV"
            ]

            joined.append(filtered_row)

        td_buckets[pod] = joined

    today = datetime.today().strftime("%m/%d/%Y")
    divider = "<hr style='border:none;border-top:1px solid #ccc;margin:20px 0;'>"
    drafted = 0

    for pod_name, pod_info in PODS.items():
        port_rows = port_buckets[pod_name]
        td_rows = td_buckets[pod_name]
        ca_rows = []

        if not port_rows and not td_rows and not ca_rows:
            continue

        sections = []

        if port_rows:
            table = _html_table(port_headers, port_rows, keep_cols=port_cols)
            sections.append(
                _section_header("Portfolio Review — Flagged Funds (NET Cash)") + table
            )

        if td_rows:
            table = _html_table(combined_td_hdrs, td_rows, uniform_width="60px")
            sections.append(
                _section_header("TD Report — Flagged Funds (Last Business Day)") + table
            )

        # if ca_rows:
        #     table = _html_table(ca_headers, ca_rows)
        #     sections.append(
        #         _section_header("Corporate Actions — Open Items") + table
        #     )

        html_body = f"""
<html>
<body style='font-family:Calibri,sans-serif;font-size:11pt;'>
<p>Hi team,</p>
<p>Please see the flagged funds for <b>{today}</b>:</p>
{divider.join(sections)}
<br>
<p>Best regards,</p>
</body>
</html>"""

        subject = f"Daily Pod Report — {today}"
        _draft_email(pod_info["to"], pod_info["cc"], subject, html_body)
        print(
            f"[OK] Draft opened : {pod_info['display']:<14} "
            f"| {len(port_rows):>2} port review flag(s)  "
            f"| {len(td_rows):>2} TD flag(s)  "
        )
        drafted += 1

    print()
    if drafted == 0:
        print("No flagged funds found — no emails drafted.")
    else:
        print(f"{drafted} draft(s) opened in Outlook.")


if __name__ == "__main__":
    main()
